# 파이썬 9일차

import openpyxl

# 기존 엑셀파일 열기 : 파일 객체명 = openpyxl.load_workbook("파일명")
# 새로운 엑셀파일 생성 : 파일 객체명 = openpyxl.Workbook()
# 현재 워크시트 선택 : 워크시트 객체명 = 파일 객체명.active
# 새로운 엑셀파일 생성 : 파일 객체명 = 파일 객체명['워크시트명']
# 파일 저장 : 파일 객체명.save("파일명")
# 파일 닫기 : 파일 객체명.close()

wb=openpyxl.load_workbook("./파일입출력/실습2.xlsx")
ws=wb.active

lotto=[]
for i in range(45) :
  lotto.append(0)

for r in range(4,895) :
  for c in range(14,20) :
    lotto[int(ws.cell(row=r,column=c).value)-1] += 1

for i in range(45) :
  print("%d번 출현 횟수 : %d"%(i+1, lotto[i]))

wb.close()