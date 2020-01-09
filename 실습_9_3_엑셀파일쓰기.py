# 파이썬 9일차

import openpyxl

wb=openpyxl.Workbook()
ws=wb.active

data=[]
sum=0

def oper() :
  a=input("수입/지출/잔액/종료 : ")
  if a=="수입" :
    func1()
    oper()
  elif a=="지출" :
    func2()
    oper()
  elif a=="잔액" :
    func3()
    oper()
  else :
    print("종료")
    ws["A1"]="수입";ws["B1"]="지출";ws["C1"]="잔액"
    for i in range(len(data)) : 
      ws.append(data[i])
    wb.save("실습3.xlsx")
    wb.close()
    exit()

def func1() :
  global sum
  a=int(input("수입 금액 : "))
  sum += a
  data.append([a,0,sum])

def func2() :
  global sum
  a=int(input("지출 금액 : "))
  sum -= a
  data.append([0,a,sum])

def func3() :
  global sum
  print("현재 잔액은 %d원입니다."%sum)

oper()