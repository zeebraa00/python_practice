# 파이썬 9일차

import openpyxl

wb=openpyxl.load_workbook("./파일입출력/실습5.xlsx")
ws=wb.active

header = [1,"P153","성균관대역"]
data=input("성균관대의 1~6월 승하차인원 : ").split()

data=list(map(int,data))
ws.append(header+data)
ws.cell(row=1,column=10).value="평균승하차인원"
for line in ws :
  for r in range(2,12) :
    sum=0
    for c in range(4,10) :
      sum+=ws.cell(row=r,column=c).value
    ws.cell(row=r,column=10).value=sum/6

sum=0
for c in range(4,10) :
  sum+=ws.cell(row=12, column=c).value
ws.cell(row=12, column=10).value=sum/6

wb.save("./파일입출력/실습5.xlsx")